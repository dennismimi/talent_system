from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse_lazy
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
import datetime as dt
import pandas as pd
from .forms import Search
import os
from openpyxl import load_workbook
from .models import Student, Fees_paid, Receipt
from django.conf import settings
from django.views.generic import DetailView, ListView
from django.core.files.storage import FileSystemStorage
from django.views.generic.edit import CreateView, DeleteView, UpdateView


def import_excel_data(request):
    try:
        if request.method == 'POST' and request.FILES['myfile']:
            file_type = request.POST['file_type']
            myfile = request.FILES['myfile']
            fs = FileSystemStorage()
            filename = fs.save(myfile.name, myfile)
            uploaded_file_url = fs.url(filename)
            excel_file = uploaded_file_url
            empexceldata = pd.read_excel("." + excel_file)
            lw_excel = load_workbook("." + excel_file)
            sheets = lw_excel.sheetnames
            dbframe = empexceldata
            if file_type == 'rct':
                for dbframe in dbframe.itertuples():
                    obj_std = get_object_or_404(Student, adm_no=dbframe.ADM_NO)
                    obj = Receipt.objects.create(std_name=obj_std.id, rct_no=dbframe.receipt_no, amount=dbframe.amt)
                    obj.save()
                    print('saved')

            elif file_type == 'std':
                for dbframe in dbframe.itertuples():
                    obj = Student.objects.create(adm_no=dbframe.ADM_NO, std_name=dbframe.NAME, std_class=dbframe.CLASS,
                                                 std_boarder=False)
                    obj.save()
                    print('saved')

            else:
                for sheet in sheets[1:]:
                    dbframe = pd.read_excel("." + excel_file, sheet_name=sheet, header=1)
                    dbframe = dbframe.fillna(0)
                    for dbframe in dbframe.itertuples():
                        obj_std = get_object_or_404(Student, adm_no=dbframe.ADM_NO)
                        obj = Fees_paid.objects.create(adm_levies=dbframe.ADM_LEVIES, uniform=dbframe.UNFM,
                                                       dev=dbframe.DEV, motiv=dbframe.MOTIV,
                                                       assesm=dbframe.ASSESM, medical=dbframe.MEDICAL,
                                                       other_levies=dbframe.OTHER_LEVIES,
                                                       bus_charge=dbframe.BUS_CH, bus_dest=dbframe.BUS_DEST,
                                                       prev_bal=dbframe.PREV_BALANCE,
                                                       term_fees=dbframe.TERM_FEES, total_amt=dbframe.TOTAL_AMOUNT,
                                                       rem_bal=dbframe.REMAINING_BALANCE,
                                                       total_amt_paid=dbframe.TOTAL_AMT_PAID,
                                                       fees_reg_bal=dbframe.FEES_REG_BALANCE, student_id=obj_std.id,
                                                       receipt_id=1)

                        obj.save()
                        print('saved')

            return render(request, 'import_excel.html', {'uploaded_file_url': uploaded_file_url})
    except Exception as identifier:
        print(identifier)

    return render(request, 'import_excel.html', {})


class StudentListView(ListView):
    template_name = "Accounts/student_list.html"
    paginate_by = 10

    def get_queryset(self):
        students = Student.objects.all()
        return students.order_by("std_name")


class ReceiptListView(ListView):
    template_name = "Accounts/receipt_list.html"
    paginate_by = 10

    def get_queryset(self):
        receipts = Receipt.objects.all()
        return receipts.order_by("date_paid")


class FeesPaidListView(ListView):
    template_name = "Accounts/feespaid_list.html"
    paginate_by = 10

    def get_queryset(self):
        feepaids = Fees_paid.objects.all()
        return feepaids.order_by("student")


def import_images(request):
    for image in images:
        pass


class CreateFeesPaidView(LoginRequiredMixin, CreateView):
    model = Fees_paid
    fields = ['adm_levies', 'uniform', 'dev', 'motiv', 'assesm',
              'medical', 'other_levies', 'bus_charge', 'bus_dest', 'prev_bal', 'fee_purpose', 'term_fees',
              'total_amt', 'rem_bal', 'total_amt_paid', 'fees_reg_bal']
    success_url = reverse_lazy('students')


class CreateStudentView(LoginRequiredMixin, CreateView):
    model = Student
    fields = ['adm_no', 'std_name', 'std_class', 'std_boarder']
    success_url = reverse_lazy('students')


class CreateReceiptView(LoginRequiredMixin, CreateView):
    model = Receipt
    fields = ['std_name', 'rct_no', 'amount', 'fee_purpose']
    success_url = reverse_lazy('students')


class UpdateFeesPaidView(LoginRequiredMixin, UpdateView):
    model = Fees_paid
    fields = ['adm_levies', 'uniform', 'dev', 'motiv', 'assesm',
              'medical', 'other_levies', 'bus_charge', 'bus_dest', 'prev_bal', 'term_fees',
              'total_amt', 'rem_bal', 'total_amt_paid', 'fees_reg_bal']
    success_url = reverse_lazy('students')


class UpdateStudentView(LoginRequiredMixin, UpdateView):
    model = Student
    fields = ['adm_no', 'std_name', 'std_class', 'std_boarder']
    success_url = reverse_lazy('students')


class UpdateReceiptView(LoginRequiredMixin, UpdateView):
    model = Receipt
    fields = ['std_name', 'rct_no', 'amount']
    success_url = reverse_lazy('students')


class DeleteFeesPaidView(LoginRequiredMixin, DeleteView):
    model = Fees_paid
    fields = ['adm_levies', 'uniform', 'dev', 'motiv', 'assesm',
              'medical', 'other_levies', 'bus_charge', 'bus_dest', 'prev_bal', 'term_fees',
              'total_amt', 'rem_bal', 'total_amt_paid', 'fees_reg_bal']
    success_url = reverse_lazy('students')


class DeleteStudentView(LoginRequiredMixin, DeleteView):
    model = Student
    fields = ['adm_no', 'std_name', 'std_class', 'std_boarder']
    success_url = reverse_lazy('students')


class DeleteReceiptView(LoginRequiredMixin, DeleteView):
    model = Receipt
    fields = ['std_name', 'rct_no', 'amount']
    success_url = reverse_lazy('students')


def searchResult(request):
    return render(request, 'Accounts/results.html')


def searchInfo(request):
    if request.method == 'POST':
        form = Search(request.POST)
        if form.is_valid():
            text_search = form.cleaned_data['text_search']
            results = Student.objects.filter(adm_no=text_search)
            if results is True:
                return render(request, 'Accounts/results.html', {'results': results})
            else:
                return render(request, 'Accounts/search_page.html')
        else:
            return render(request, 'Accounts/search_page.html')

    return render(request, 'Accounts/search_page.html')